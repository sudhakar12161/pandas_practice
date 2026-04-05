from __future__ import annotations

from datetime import datetime
import logging
from pathlib import Path
import re
import time
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from pyspark import StorageLevel
from pyspark.sql import functions as F

from app.services.datasource_loader import DataSourceLoader


class CompareService:
    """Schema compare, profile metrics, and row compare operations."""

    def __init__(
        self,
        loader: DataSourceLoader,
        export_dir: Path,
        max_error_rows: int,
        repartition_threshold: int = 500_000,
    ) -> None:
        self.logger = logging.getLogger("data_compare.compare.pyspark")
        self.loader = loader
        self.export_dir = export_dir
        self.max_error_rows = max_error_rows
        self.repartition_threshold = repartition_threshold
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def resolve_schema(self, side: dict[str, Any], saved_schema: dict[str, Any] | None) -> dict[str, str]:
        start = time.perf_counter()
        schema_name = side.get("schema_name")
        if schema_name and schema_name not in {"NONE", "INFER_SCHEMA"} and saved_schema:
            schema = self._to_schema_map(saved_schema["schema_def"])
            self.logger.info("Resolve schema from saved schema_name=%s columns=%s elapsed_ms=%.1f", schema_name, len(schema), (time.perf_counter() - start) * 1000)
            return schema
        if schema_name == "INFER_SCHEMA":
            schema = self._to_schema_map(self.loader.infer_schema(side))
            self.logger.info("Resolve schema inferred columns=%s elapsed_ms=%.1f", len(schema), (time.perf_counter() - start) * 1000)
            return schema
        df = self.loader.load_dataframe(side)
        schema = {f.name: f.dataType.simpleString() for f in df.schema.fields}
        self.logger.info("Resolve schema from dataframe columns=%s elapsed_ms=%.1f", len(schema), (time.perf_counter() - start) * 1000)
        return schema

    def compare_schema(self, source: dict[str, str], target: dict[str, str]) -> list[dict[str, Any]]:
        source_map = self._to_schema_map(source)
        target_map = self._to_schema_map(target)
        all_columns = sorted(set(source_map).union(set(target_map)))

        result = []
        for col in all_columns:
            src_type = source_map.get(col)
            tgt_type = target_map.get(col)
            result.append(
                {
                    "source_column": col if col in source_map else None,
                    "source_type": src_type,
                    "target_column": col if col in target_map else None,
                    "target_type": tgt_type,
                    "match": src_type == tgt_type and src_type is not None,
                }
            )
        return result

    def analyze(self, side: dict[str, Any], exclude_columns: list[str], pk_columns: list[str]) -> dict[str, Any]:
        start = time.perf_counter()
        df = self.loader.load_dataframe(side)
        columns = [col for col in df.columns if col not in exclude_columns]
        total_rows = df.count()
        metric_rows: list[dict[str, Any]] = []

        for metric_name in ["NOT_NULL_COUNT", "NULL_COUNT", "UNIQUE_COUNT", "DUPLICATE_COUNT", "MIN", "MAX"]:
            metric_rows.append({"metric": metric_name, "values": {}})

        def set_metric(metric_name: str, col_name: str, value: Any) -> None:
            row = next(item for item in metric_rows if item["metric"] == metric_name)
            row["values"][col_name] = value

        for col_name in columns:
            null_count = df.filter(F.col(col_name).isNull()).count()
            not_null_count = total_rows - null_count
            unique_count = df.select(F.col(col_name)).distinct().count()
            duplicate_count = total_rows - unique_count
            min_max = df.agg(F.min(F.col(col_name)).alias("min_val"), F.max(F.col(col_name)).alias("max_val")).first()

            set_metric("NOT_NULL_COUNT", col_name, int(not_null_count))
            set_metric("NULL_COUNT", col_name, int(null_count))
            set_metric("UNIQUE_COUNT", col_name, int(unique_count))
            set_metric("DUPLICATE_COUNT", col_name, int(duplicate_count))
            set_metric("MIN", col_name, str(min_max["min_val"]))
            set_metric("MAX", col_name, str(min_max["max_val"]))

        result = {
            "dataset_name": self.get_dataset_name(side),
            "columns": columns,
            "rows": metric_rows,
            "pk_columns": pk_columns,
        }
        self.logger.info(
            "Analyze completed dataset=%s columns=%s rows=%s elapsed_ms=%.1f",
            result["dataset_name"],
            len(columns),
            total_rows,
            (time.perf_counter() - start) * 1000,
        )
        return result

    def export_analysis_excel(self, source_stats: dict[str, Any], target_stats: dict[str, Any]) -> str:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_path = self.export_dir / f"analyze_report_{timestamp}.xlsx"
        source_df = self._analysis_table_df(source_stats)
        target_df = self._analysis_table_df(target_stats)

        wb = Workbook()
        ws_source = wb.active
        ws_source.title = "Source Data Statistics"
        ws_target = wb.create_sheet("Target Data Statistics")

        self._write_df_with_light_table(ws_source, source_df, "SourceStatsTable")
        self._write_df_with_light_table(ws_target, target_df, "TargetStatsTable")

        wb.save(file_path)

        return str(file_path)

    def compare_data(
        self,
        source_side: dict[str, Any],
        target_side: dict[str, Any],
        source_pk_columns: list[str],
        target_pk_columns: list[str],
        column_mapping: dict[str, str],
    ) -> dict[str, Any]:
        start = time.perf_counter()
        source_df = self.loader.load_dataframe(source_side)
        target_df = self.loader.load_dataframe(target_side)

        if len(source_pk_columns) != len(target_pk_columns):
            raise ValueError("Primary key column counts must match between Source and Target")

        pk_pairs = []
        for src_pk in source_pk_columns:
            tgt_pk = column_mapping.get(src_pk)
            if not tgt_pk:
                raise ValueError(f"Missing mapping for source PK column: {src_pk}")
            pk_pairs.append((src_pk, tgt_pk))

        source_pk_cols = [src for src, _ in pk_pairs]
        target_pk_cols = [tgt for _, tgt in pk_pairs]

        if len(set(target_pk_cols)) != len(target_pk_cols):
            raise ValueError("Each source PK must map to a unique target PK")
        if set(target_pk_cols) != set(target_pk_columns):
            raise ValueError("Mapped target PKs must exactly match selected target PK columns")

        missing_source = [col for col in source_pk_cols if col not in source_df.columns]
        missing_target = [col for col in target_pk_cols if col not in target_df.columns]
        if missing_source or missing_target:
            raise ValueError(
                f"PK columns missing. source_missing={missing_source}, target_missing={missing_target}"
            )

        needed_source_cols = sorted(set(source_pk_cols + [col for col in column_mapping.keys() if col in source_df.columns]))
        needed_target_cols = sorted(set(target_pk_cols + [col for col in column_mapping.values() if col in target_df.columns]))
        source_df = source_df.select(*needed_source_cols)
        target_df = target_df.select(*needed_target_cols)

        source_total = source_df.count()
        target_total = target_df.count()
        source_unique = source_df.dropDuplicates(source_pk_cols).count()
        target_unique = target_df.dropDuplicates(target_pk_cols).count()

        if max(source_total, target_total) >= self.repartition_threshold:
            source_df = source_df.repartition(*[F.col(col) for col in source_pk_cols])
            target_df = target_df.repartition(*[F.col(col) for col in target_pk_cols])

        source_alias = source_df.alias("s")
        target_alias = target_df.alias("t")
        # Compare PKs as normalized strings to avoid cross-source type mismatch
        # (for example file string vs database numeric/date).
        join_expr = [
            (
                F.col(f"s.{src_pk}").isNotNull()
                & F.col(f"t.{tgt_pk}").isNotNull()
                & (self._norm_expr_no_null(F.col(f"s.{src_pk}")) == self._norm_expr_no_null(F.col(f"t.{tgt_pk}")))
            )
            for src_pk, tgt_pk in pk_pairs
        ]
        joined = source_alias.join(target_alias, join_expr, "inner").persist(StorageLevel.MEMORY_AND_DISK)

        matched_rows = joined.count()
        rows_only_in_source = source_alias.join(target_alias, join_expr, "left_anti").count()
        rows_only_in_target = target_alias.join(source_alias, join_expr, "left_anti").count()
        not_matched_rows = rows_only_in_source + rows_only_in_target
        column_stats = {}
        mismatched_columns: set[str] = set()
        mismatch_expr = F.lit(False)
        for source_col, target_col in column_mapping.items():
            if source_col in source_pk_cols:
                continue
            if source_col not in source_df.columns or target_col not in target_df.columns:
                continue
            compare_expr = self._norm_expr(F.col(f"s.{source_col}")) == self._norm_expr(F.col(f"t.{target_col}"))
            matches = joined.filter(compare_expr).count()
            mismatches = matched_rows - matches
            column_stats[source_col] = {
                "target_column": target_col,
                "matching_count": matches,
                "non_matching_count": mismatches,
            }
            if mismatches > 0:
                mismatched_columns.add(source_col)
            mismatch_expr = mismatch_expr | (~compare_expr)

        mismatch_rows = joined.filter(mismatch_expr)
        mismatch_count = mismatch_rows.count()
        data_payload = self._build_data_tab_payload(
            source_alias=source_alias,
            target_alias=target_alias,
            source_df=source_df,
            target_df=target_df,
            join_expr=join_expr,
            source_pk_cols=source_pk_cols,
            target_pk_cols=target_pk_cols,
            column_mapping=column_mapping,
        )

        source_name = self.get_dataset_name(source_side)
        target_name = self.get_dataset_name(target_side)
        common_name = "COMMON"
        mismatch_columns_text = ", ".join(sorted(mismatched_columns)) if mismatched_columns else ""

        report = {
            "source": {
                "dataset_name": source_name,
                "total_rows": source_total,
                "unique_rows_by_pk": source_unique,
                "duplicate_rows_by_pk": source_total - source_unique,
                "rows_only_in_source": rows_only_in_source,
            },
            "target": {
                "dataset_name": target_name,
                "total_rows": target_total,
                "unique_rows_by_pk": target_unique,
                "duplicate_rows_by_pk": target_total - target_unique,
                "rows_only_in_target": rows_only_in_target,
            },
            "common": {
                "dataset_name": common_name,
                "matched_rows_by_pk": matched_rows,
                "not_matched_rows_by_pk": not_matched_rows,
                "mismatched_columns": mismatch_columns_text,
                "column_level_stats": column_stats,
                "mismatch_row_count": mismatch_count,
            },
        }
        report["excel_path"] = self._write_excel(report, data_payload, source_name)
        joined.unpersist()
        self.logger.info(
            "Compare completed source=%s target=%s matched=%s mismatched_rows=%s elapsed_ms=%.1f",
            source_name,
            target_name,
            matched_rows,
            mismatch_count,
            (time.perf_counter() - start) * 1000,
        )
        return report

    def _write_excel(self, report: dict[str, Any], data_payload: dict[str, Any], source_name: str) -> str:
        date_stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        safe_source_name = re.sub(r"[^\w\-]+", "_", source_name).strip("_") or "source"
        file_path = self.export_dir / f"{safe_source_name}_{date_stamp}.xlsx"

        summary_rows = []
        for section_key in ["source", "target", "common"]:
            values = report[section_key]
            dataset_name = values.get("dataset_name", section_key.upper())
            for key, value in values.items():
                if key in {"dataset_name", "column_level_stats", "mismatch_row_count"}:
                    continue
                summary_rows.append(
                    {
                        "DATASET_NAME": dataset_name,
                        "METRIC_NAME": key.upper(),
                        "METRIC_VALUE": value,
                    }
                )

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["DATASET_NAME", "METRIC_NAME", "METRIC_VALUE"])
        for row in summary_rows:
            ws_summary.append([row["DATASET_NAME"], row["METRIC_NAME"], row["METRIC_VALUE"]])
        summary_last_row = ws_summary.max_row
        summary_table = Table(displayName="SummaryTable", ref=f"A1:C{summary_last_row}")
        summary_style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        summary_table.tableStyleInfo = summary_style
        ws_summary.add_table(summary_table)

        ws_data = wb.create_sheet("Data")
        ws_data.append(data_payload["headers"])
        red_font = Font(color="00FF0000")
        thin = Side(style="thin", color="00D9D9D9")
        all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(data_payload["headers"]) + 1):
            ws_data.cell(row=1, column=col_idx).border = all_border

        for row_idx, row_item in enumerate(data_payload["rows"], start=2):
            values = row_item["values"]
            mismatch_indices = row_item["mismatch_indices"]
            ws_data.append(values)
            for col_idx in range(1, len(values) + 1):
                ws_data.cell(row=row_idx, column=col_idx).border = all_border
            if row_item.get("matched_ind") == "Mismatched":
                for col_idx in mismatch_indices:
                    ws_data.cell(row=row_idx, column=col_idx + 1).font = red_font

        wb.save(file_path)

        return str(file_path)

    def _analysis_table_df(self, stats: dict[str, Any]) -> pd.DataFrame:
        rows = []
        for metric_row in stats["rows"]:
            row = {"METRIC": metric_row["metric"]}
            row.update(metric_row["values"])
            rows.append(row)
        return pd.DataFrame(rows)

    def _write_df_with_light_table(self, worksheet, dataframe: pd.DataFrame, table_name: str) -> None:
        worksheet.append(list(dataframe.columns))
        for _, row in dataframe.iterrows():
            worksheet.append(list(row.values))

        last_row = worksheet.max_row
        last_col = worksheet.max_column
        if last_row < 1 or last_col < 1:
            return
        end_col = self._excel_col_name(last_col)
        table_ref = f"A1:{end_col}{last_row}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)

    def _excel_col_name(self, index: int) -> str:
        result = ""
        while index:
            index, rem = divmod(index - 1, 26)
            result = chr(65 + rem) + result
        return result

    def get_dataset_name(self, side: dict[str, Any]) -> str:
        if side.get("uploaded_name"):
            return str(side["uploaded_name"])
        if side.get("table_name"):
            return side["table_name"]
        if side.get("uploaded_path"):
            file_name = Path(side["uploaded_path"]).name
            stripped = re.sub(r"^[a-f0-9]{32}_", "", file_name)
            return Path(stripped).stem
        if side.get("file_path"):
            return Path(side["file_path"]).stem
        connection = side.get("connection") or {}
        return connection.get("name", "DATASET")

    def _to_schema_map(self, schema: Any) -> dict[str, str]:
        if isinstance(schema, dict):
            return {str(k): str(v) for k, v in schema.items()}
        if isinstance(schema, list):
            out: dict[str, str] = {}
            for item in schema:
                if isinstance(item, dict) and item.get("name"):
                    out[str(item["name"])] = str(item.get("type", "string"))
            return out
        return {}

    def _build_data_tab_payload(
        self,
        source_alias,
        target_alias,
        source_df,
        target_df,
        join_expr,
        source_pk_cols: list[str],
        target_pk_cols: list[str],
        column_mapping: dict[str, str],
    ) -> dict[str, Any]:
        full_join = source_alias.join(target_alias, join_expr, "full_outer")

        source_present_expr = F.lit(True)
        for col in source_pk_cols:
            source_present_expr = source_present_expr & F.col(f"s.{col}").isNotNull()

        target_present_expr = F.lit(True)
        for col in target_pk_cols:
            target_present_expr = target_present_expr & F.col(f"t.{col}").isNotNull()

        headers: list[str] = []
        select_exprs = []
        pair_indices: list[tuple[int, int]] = []

        # PK coalesced columns first.
        for src_pk, tgt_pk in zip(source_pk_cols, target_pk_cols):
            alias = f"PK_{src_pk}" if src_pk == tgt_pk else f"PK_{src_pk}_{tgt_pk}"
            headers.append(alias)
            select_exprs.append(F.coalesce(F.col(f"s.{src_pk}"), F.col(f"t.{tgt_pk}")).alias(alias))

        mismatch_expr = F.lit(False)
        non_pk_pairs: list[tuple[str, str]] = []
        for source_col, target_col in column_mapping.items():
            if source_col in source_pk_cols:
                continue
            s_name = f"S_{source_col}"
            t_name = f"T_{target_col}"
            if source_col not in source_df.columns or target_col not in target_df.columns:
                continue
            non_pk_pairs.append((source_col, target_col))
            src_alias = f"SRC_{source_col}"
            tgt_alias = f"TGT_{target_col}"
            src_idx = len(headers)
            headers.extend([src_alias, tgt_alias])
            pair_indices.append((src_idx, src_idx + 1))
            select_exprs.append(F.col(f"s.{source_col}").alias(src_alias))
            select_exprs.append(F.col(f"t.{target_col}").alias(tgt_alias))
            mismatch_expr = mismatch_expr | (
                self._norm_expr(F.col(f"s.{source_col}")) != self._norm_expr(F.col(f"t.{target_col}"))
            )

        status_col = (
            F.when(source_present_expr & (~target_present_expr), F.lit("Only in Source"))
            .when((~source_present_expr) & target_present_expr, F.lit("Only in Target"))
            .when(source_present_expr & target_present_expr & mismatch_expr, F.lit("Mismatched"))
            .otherwise(F.lit("Matched"))
        )

        headers.append("Matched_Ind")
        selected_df = full_join.select(*select_exprs, status_col.alias("Matched_Ind")).limit(self.max_error_rows)

        rows_payload: list[dict[str, Any]] = []
        for row in selected_df.collect():
            values = [row[h] for h in headers]
            mismatch_indices: set[int] = set()
            for src_idx, tgt_idx in pair_indices:
                src_val = values[src_idx]
                tgt_val = values[tgt_idx]
                if self._norm(src_val) != self._norm(tgt_val):
                    mismatch_indices.add(src_idx)
                    mismatch_indices.add(tgt_idx)
            rows_payload.append(
                {
                    "values": values,
                    "mismatch_indices": sorted(mismatch_indices),
                    "matched_ind": values[-1],
                }
            )

        return {"headers": headers, "rows": rows_payload}

    def _norm(self, value: Any) -> str:
        return "__NULL__" if value is None else str(value).strip()

    def _norm_expr(self, expr):
        # Trim string representation to reduce false mismatches from whitespace.
        return F.coalesce(F.trim(expr.cast("string")), F.lit("__NULL__"))

    def _norm_expr_no_null(self, expr):
        return F.trim(expr.cast("string"))
